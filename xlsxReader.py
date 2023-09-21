import openpyxl
from shiftAssigner import Person

def xlsxReader(xlsxFileName):
    # load excel with its path
    workbook = openpyxl.load_workbook(xlsxFileName)

    sh = workbook.active
    
    people = []

    # iterate through excel and display data
    for i in range(2, sh.max_row+1):

        name = ""
        shifts = []
        unavailableShifts = []
        
        for j in range(6, sh.max_column+1):
            cellData = sh.cell(row=i, column=j)
            
            match j:
                # name
                case 6:
                    name = cellData.value
                
                # unavailable shifts
                case 12:
                    if (cellData.value != "nil" and cellData.value != None):
                        unavailableShiftsStr = str(cellData.value).split(",")
                        unavailableShifts = [int(shift) for shift in unavailableShiftsStr]
                
                # monday - fridays
                case default:
                    if(cellData.value != "nil" and cellData.value != None):
                        shifts.extend(convertShiftNumber(j, cellData.value))
        
        personInstance = Person(name, shifts, unavailableShifts)
        people.append(personInstance)
    return people
    
# Gives a (day, shift) where day: 0-4 for mon-fri and shift: 0-2 for order of shift in a day
def convertShiftNumber(column, shiftsSelected):
    
    shiftsSelected = shiftsSelected.split(";")
    convertedShifts = []
    
    for shift in shiftsSelected:
        match shift:
            case "930am-1230pm":
                convertedShifts.append( (column-7, 0) )
            
            case "1230pm-330pm":
                convertedShifts.append( (column-7, 1) )

            case "330pm-630pm":
                convertedShifts.append( (column-7, 2) )
                
    return convertedShifts
