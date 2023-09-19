# import module
import openpyxl

# load excel with its path
workbook = openpyxl.load_workbook("C:\\Users\\ASUS\\Downloads\\Test for SMU Shop(1-4).xlsx")

sh = workbook.active
dataDict = {}

# iterate through excel and display data
for i in range(1, sh.max_row+1):
	print("\n")
	print("Row ", i, " data :")
	
	for j in range(1, sh.max_column+1):
		cell_obj = sh.cell(row=i, column=j)
		print(cell_obj.value, end=" ")
